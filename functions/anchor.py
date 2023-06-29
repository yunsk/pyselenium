
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.common.exceptions import WebDriverException,NoSuchElementException 
from webdriver_manager.chrome import ChromeDriverManager
from urllib.error import URLError, HTTPError
import datetime,time, os

class BasePage(object):
    def __init__(self, driver):
        self.driver = driver

class MainFunction(BasePage):
    def anchorTest(self,excelName,sheetname,windowsize) :
        self.driver.set_window_size(*windowsize) # 가로 세로 크기를 언팩킹하여 전달
        self.wb = load_workbook(filename=excelName)
        self.ws = self.wb[sheetname]
        start = time.time()

        for i in range(2, self.ws.max_row+1) :
            no =  self.ws['A'+str(i)].value
            anchorId = self.ws['B'+str(i)].value
            global_url = self.ws['C'+str(i)].value
            try: 
                self.driver.get(global_url)
                time.sleep(2)
                
                # 타겟 경로 구함
                folder_name = "anchor_Screenshot"
                current_directory = os.getcwd()
                new_folder_path = os.path.join(current_directory,folder_name)
                if not os.path.exists(new_folder_path):
                    os.makedirs(new_folder_path)
                else:
                    pass
                #스샷 저장
                self.driver.save_screenshot(folder_name+'\\'+anchorId+".png")
                time.sleep(2) 
                #driver.execute_script("window.scrollTo(0,22145)")
            
            except NoSuchElementException  :
                print('찾은 엘레먼트가 없음')
            except WebDriverException:
                print('Driver ERROR')
            except HTTPError as e : 
                print("ERROR"+str(e.code))
            except URLError as e :
                print("reason"+str(e.reason)) 
            
        #driver.close()
        self.wb.save('Result_AnchorID'+'.xlsx')
        sec = time.time()-start # 종료
        times = str(datetime.timedelta(seconds=sec))
        short = times.split(".")[0] #초단위 

        print('완료시간',f"{short} sec")
