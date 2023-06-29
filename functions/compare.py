
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.common.exceptions import WebDriverException,NoSuchElementException 
from webdriver_manager.chrome import ChromeDriverManager
from urllib.error import URLError, HTTPError
import datetime,time, os

class BasePage(object):
    def __init__(self, driver):
        self.driver = driver

class MainFunction(BasePage):
    def CompareTest(self,excelName,windowsize) :
        self.wb = load_workbook(filename=excelName)
        self.ws = self.wb.active
        self.driver.set_window_size(*windowsize)
        start = time.time() #시작

        TestURL = self.driver.get('https://www.samsung.com/global/galaxy/galaxy-s23/compare/')

        time.sleep(2)
        #document.
        self.driver.execute_script("window.scrollTo(0,307)")

        model_sum = self.driver.find_elements(By.XPATH,'//*[@id="model-colors"]/ul[1]/li[1]/div[1]/ul/li')
        print('모델총 갯수 : ',len(model_sum))

        ex_cell = 2 

        try : 
            for n in range(1,len(model_sum)+1) : 
                #Excel 첫번째 행에 no 입력 
                self.ws['A'+str(ex_cell)] = n

                model_xpath = '//*[@id="model-colors"]/ul[1]/li[1]/div[1]/ul/li[{0}]'.format(n)
                model_name = self.driver.find_element(By.XPATH,model_xpath).text
                select_model_xpath = self.driver.find_element(By.CLASS_NAME,'select-device')

                #드롭다운 메뉴 
                select_model = self.driver.find_element(By.CLASS_NAME,'select-device').text.replace('\n','')
                #colors
                colors = self.driver.find_elements(By.XPATH,'//*[@id="model-colors"]/ul[1]/li[1]/div[3]/div[1]/ul/li')
                color_num = self.driver.find_element(By.XPATH,'//*[@id="model-colors"]/ul[1]/li[1]/div[3]/div[1]/ul/li').text
                #exclusiveColor
                color_exclusive = self.driver.find_elements(By.XPATH,'//*[@id="model-colors"]/ul[1]/li[1]/div[3]/div[2]/ul/li')
                exclusive_name = self.driver.find_element(By.XPATH,'//*[@id="model-colors"]/ul[1]/li[1]/div[3]/div[2]/ul/li').text
                print(f'n몇번째:',n)
                #CTA
                buy_cta = self.driver.find_element(By.XPATH,'//*[@id="model-colors"]/ul[1]/li[1]/div[4]/div[1]/a[1]').text
                print('buycta존재',buy_cta,len(buy_cta))
                see_cta = self.driver.find_element(By.XPATH,'//*[@id="model-colors"]/ul[1]/li[1]/div[4]/div[2]/button').text
                print('seecta존재',see_cta,len(see_cta))
                
                #시작 값 보다 큰 경우 다음 모델 이어서 진행 
                if(n>1):
                    select_model_xpath.click()
                    time.sleep(1)
                    self.driver.find_element(By.XPATH,model_xpath).click()
                    time.sleep(1)
                    select_model_xpath = self.driver.find_element(By.CLASS_NAME,'select-device').text.replace('\n','')
                
                if len(buy_cta) == 0 : 
                    self.ws.cell(row=ex_cell,column=8).value="CTA없음"
                elif len(see_cta) == 0 :
                    self.ws.cell(row=ex_cell,column=9).value="CTA없음"
                else : 
                    print('입력n',ex_cell)
                    self.ws.cell(row=ex_cell,column=8).value=buy_cta
                    self.ws.cell(row=ex_cell,column=9).value=see_cta
                

                select_color1 = self.driver.find_element(By.XPATH,'//*[@id="model-colors"]/ul[1]/li[1]/div[3]/div[1]/p[2]').text
                print(select_color1)
                #colors_num
                color_num = self.driver.find_element(By.XPATH,'//*[@id="model-colors"]/ul[1]/li[1]/div[3]/div[1]/ul/li').text
                print(color_num)
                #exclusive_num  
                exclusive_num = self.driver.find_element(By.XPATH,'//*[@id="model-colors"]/ul[1]/li[1]/div[3]/div[2]/ul/li').text
                print(exclusive_num)
            
                #스크린샷 폴더 설정
                folder_name = "Compare_ScrennShot"
                current_directory = os.getcwd()
                new_folder_path = os.path.join(current_directory,folder_name)
                if not os.path.exists(new_folder_path):
                    os.makedirs(new_folder_path)
                else:
                    pass

            #Color칩 선택 
                for i in colors : 
                    if i.is_displayed()==True : 

                        i.click()
                        time.sleep(2)
                        colors_name= self.driver.find_element(By.XPATH,'//*[@id="model-colors"]/ul[1]/li[1]/div[3]/div[1]/p[2]').text.replace('\n','')
                        print('선택한',select_model+'_cell_',str(ex_cell)+'colors_Name=',colors_name)
                    if len(colors_name)== 0 :
                        self.ws['E'+str(ex_cell)] = '없음'
                    else : 
                        self.ws['B'+str(ex_cell)] = select_model
                        print('B',ex_cell,'=',select_model)
                        self.ws['C'+str(ex_cell)] = colors_name
                        print('C',ex_cell,'=',colors_name)
                        
                        model_img_xpath = self.driver.find_element(By.XPATH,'//*[@id="model-colors"]/ul[1]/li[1]/div[2]/img')
                        model_img_xpath.screenshot(folder_name+'\\'+select_model+'_'+colors_name+".png")

                        model_img = Image(folder_name+'\\'+select_model+'_'+colors_name+".png")
                        model_img.height =140
                        model_img.width =140
                        self.ws.add_image(model_img,'E'+str(ex_cell))

                        img_model_alt=model_img_xpath.get_attribute('alt')
                        img_model_src=model_img_xpath.get_attribute('src')
                        if len(img_model_alt) == 0 :
                            self.ws['F'+str(ex_cell)] = "image alt값 없음"
                        elif len(img_model_src) == 0 :
                            self.ws['G'+str(ex_cell)] = "image src값 없음"
                        else : 
                            self.ws['F'+str(ex_cell)] = img_model_alt.strip()
                            self.ws['G'+str(ex_cell)] = img_model_src.strip()
                        ex_cell=ex_cell+1
                        print(ex_cell)

                #exColor칩 선택 
                for j in color_exclusive :
                    if j.is_displayed()==True :
                        
                        j.click()
                        time.sleep(2)
                        exclusive_name = self.driver.find_element(By.XPATH,'//*[@id="model-colors"]/ul[1]/li[1]/div[3]/div[2]/p[2]').text.replace('\n','')
                        if len(exclusive_name)== 0 :
                            self.ws['E'+str(ex_cell)] = '없음'
                        else : 
                            self.ws['B'+str(ex_cell)] = select_model
                            self.ws['D'+str(ex_cell)] = exclusive_name    
                            
                            model_img_xpath.screenshot(folder_name+'\\'+select_model+'_'+exclusive_name+".png")
                            img_model_alt=model_img_xpath.get_attribute('alt')
                            img_model_src=model_img_xpath.get_attribute('src')
                            time.sleep(2)
                            model_img2 = Image(folder_name+'\\'+select_model+'_'+exclusive_name+".png")
                            model_img2.height =140
                            model_img2.width =140
                            self.ws.add_image(model_img2,'E'+str(ex_cell))

                            if len(img_model_alt) == 0 :
                                self.ws['F'+str(ex_cell)] = "image alt값 없음"
                            elif len(img_model_src) == 0 :
                                self.ws['G'+str(ex_cell)] = "image src값 없음"
                            else : 
                                self.ws['F'+str(ex_cell)] = img_model_alt
                                self.ws['G'+str(ex_cell)] = img_model_src
                            print('ex',ex_cell)
                            ex_cell=ex_cell+1
        except NoSuchElementException  :
            print('찾은 엘레먼트가 없음')
        except WebDriverException:
            print('Driver ERROR')
        except HTTPError as e : 
            print("ERROR"+str(e.code))
        except URLError as e :
            print("reason"+str(e.reason)) 

        self.wb.save('Result_Compare'+'.xlsx')
        sec = time.time()-start # 종료
        times = str(datetime.timedelta(seconds=sec))
        short = times.split(".")[0] #초단위 

        print('완료시간',f"{short} sec")