
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.common.exceptions import WebDriverException,NoSuchElementException 
from webdriver_manager.chrome import ChromeDriverManager
from urllib.error import URLError, HTTPError
import datetime,time, os

class BasePage(object):
    def __init__(self, driver):
        self.driver = driver

class MainFunction(BasePage):
    def kvCTATest(self,excelName) :
                
        wb = load_workbook(excelName)
        ws = wb.active
        start = time.time() # 시작

        for i in range(2, ws.max_row+1) :
            country = ws['B'+str(i)].value
            country_url = ws['C'+str(i)].value
            print(i-1,country,country_url)
            try: 
                self.driver.maximize_window()
                self.driver.get(country_url)
                time.sleep(2)

                #쿠키 바 닫기
                self.driver.execute_script("$('.cookie-bar__close').click()")
                self.driver.execute_script("$('.truste-button1').click()")
                
                # 타겟 경로 구함
                kv_div = self.driver.find_element(By.CSS_SELECTOR,'.highlights-kv__text')
                kv_div_a = self.driver.find_element(By.CSS_SELECTOR,'.highlights-kv__text a')
                
                banner_div = self.driver.find_element(By.CSS_SELECTOR,'#contents > div.common-banner > div > div.common-banner__item.common-banner__buynow > div > div.common-banner__text > div > div')
                banner_div_a = self.driver.find_element(By.CSS_SELECTOR,'#contents > div.common-banner > div > div.common-banner__item.common-banner__buynow > div > div.common-banner__text > div > div > a')
            
                folder_name = "KVCTA_Screenshot"
                current_directory = os.getcwd()
                new_folder_path = os.path.join(current_directory,folder_name)
                if not os.path.exists(new_folder_path):
                    os.makedirs(new_folder_path)
                else:
                    pass

                kv_shot = kv_div.screenshot(folder_name+'\\'+'KV_'+country+"_D3.png")
                
                kv_cta = kv_div_a.text
                print('KV CTA 이름'+kv_cta)
                banner_cta = banner_div.text
                print('banner_cta 이름 ',banner_cta)
                kv_url = kv_div_a.get_attribute('href')
                banner_url = banner_div_a.get_attribute('href')
            
                #스샷 저장
                time.sleep(3) 
                self.driver.execute_script("window.scrollTo(0,22145)")
            

                banner_shot = banner_div.screenshot(folder_name+'\\'+'banner_'+country+"_D3.png")
            
                if kv_cta is not None: 
                    print('kv_ctaurl',kv_url)
                    ws['D'+str(i)] = kv_cta   
                    ws['E'+str(i)] = kv_url 
                    kv_img = Image(folder_name+'\\'+'KV_'+country+"_D3.png")
                    ws.add_image(kv_img, 'L'+str(i))
                    time.sleep(1)
                else :
                    print('CTA 없음')
                    ws['D'+str(i)] = "비노출"
                    time.sleep(1)

                if banner_cta is not None: 
                    print('banner_ctaurl',banner_url)
                    ws['H'+str(i)] = banner_cta   
                    ws['I'+str(i)] = banner_url 
                    banner_img = Image(folder_name+'\\'+'banner_'+country+"_D3.png")
                    banner_img.height =90
                    banner_img.width =240
                    ws.add_image(banner_img, 'M'+str(i))
                    time.sleep(1)
                else :
                    print('banner CTA 없음')
                    ws['H'+str(i)] = "비노출"
                    time.sleep(1)
            except NoSuchElementException  :
                print('찾은 엘레먼트가 없음')
            except WebDriverException:
                print('Driver ERROR')
            except HTTPError as e : 
                print("ERROR"+str(e.code))
            except URLError as e :
                print("reason"+str(e.reason)) 
            
        wb.save('Result_KV_CTA(D3)'+'.xlsx')
        sec = time.time()-start # 종료
        times = str(datetime.timedelta(seconds=sec))
        short = times.split(".")[0] #초단위 

        print('완료시간',f"{short} sec")